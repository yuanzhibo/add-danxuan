import requests
import json
from openpyxl import load_workbook

from util import Properties

props = Properties('add-danxuan.ini').getProperties()
print(props['ip'])

# 登录
net_url = 'http://' + props['ip'] + ':8000'
login_endpoint = props['login_endpoint']
excel_file = "danxuan.xlsx"
sheet_name = "工作表1"
headers = {'Content-Type': 'application/json', 'Connection': 'keep-alive'}

workbook = load_workbook(filename=excel_file)
sheet = workbook[sheet_name]
session = requests.Session()

response = session.post(f'{net_url}{login_endpoint}', headers=headers, data=json.dumps({
    "userName": props['usr'],
    "password": props['pwd'],
    "remember": True
}))
print(response.text)

for row in sheet.iter_rows(min_row=2, values_only=True):
    print(str(row[3]) + ' -- is adding...')
    questionType = row[0]
    gradeLevel = row[1]
    subjectId = row[2]
    title = row[3]
    options_str = row[4]
    analyze = row[5]
    correct = row[6]
    score = row[7]
    difficult = row[8]
    # 拆分字符串得到选项列表
    options_list = options_str.split(';')
    json_data = []
    for i, option in enumerate(options_list, start=1):
        json_obj = {
            "prefix": chr(64 + i),
            "content": option.strip()
        }
        json_data.append(json_obj)

    payload = {
        "id": None,
        "questionType": questionType,
        "gradeLevel": gradeLevel,
        "subjectId": subjectId,
        "title": title,
        "items": json_data,
        "analyze": analyze,
        "correct": correct,
        "score": score,
        "difficult": row[8]
    }
    
    response = session.post(net_url + '/api/admin/question/edit', headers=headers, data=json.dumps(payload))
    if response.status_code < 300:
        data = json.dumps(response.json(), indent=4)
        print(data)
    else:
        print(f"失败，状态码：{response.status_code}")
        print(response.text)

workbook.close()
